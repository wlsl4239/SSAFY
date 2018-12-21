import urllib.request

from bs4 import BeautifulSoup
from openpyxl import Workbook
from elice_utils import EliceUtils

elice_utils = EliceUtils()


def main():
    wb = Workbook()

    sheet1 = wb.active
    file_name = 'sample.xlsx'

    url = "http://www.newsis.com/eco/list/?cid=10400&scid=10404"
    req = urllib.request.Request(url)
    sourcecode = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(sourcecode, "html.parser")

    articles = []
    

    for i in soup.find_all("strong", class_="title"):
        articles.append(i.get_text())
        
    for row_index in range(1,len(articles)+1):
        sheet1.cell(row=row_index, column=1).value = row_index
        sheet1.cell(row=row_index, column=2).value = articles[row_index-1]

    wb.save(filename=file_name)
    elice_utils.send_file(file_name)


if __name__ == "__main__":
    main()

'''
엑셀 시트로 저장하기 
실생활에서는 정보를 엑셀에서 다루는 경우가 많습니다.

웹 상에서 크롤링해 온 정보를

openpyxl 라이브러리를 사용해서

엑셀 row, column 위치를 지정해주고 알맞게 저장하는 방법을 실습해봅니다.
'''
